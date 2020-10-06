#########################################################
# Title: StringFinderInExcel.py
# Author: Scott Bossard
# Date: 10/06/2020
# Description: Finds all instancs of a filename (.cpp, .h) in a excel spreadsheet and outputs them into a new spreadsheet.
##########################################################

import openpyxl

def main():
    """Main function which reads and writes files. """
    prompt = '> '

    print ("Ready to search!")
    print ("Make sure openpyxl is intalled. (> pip install openpyxl)")
    print ("Enter file path of .xlsx file: ")

    filename = input(prompt) #file path of xlsx containing data to be searched

    wb = openpyxl.load_workbook(filename) # Open file
    ws = wb.active # Open sheet

    rows = (ws.max_row) - 1 #Subtract header
    columns = ws.max_column

    print ("Searching through " + str(rows) + " rows and "+ str(columns) + " columns... ")

    filenameArray = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                if type(cell.value) != int:
                    if '.cpp' in cell.value or '.h' in cell.value:
                        arr = cell.value.split(' ')
                        matchers = ['cpp','.h']
                        matching = [s for s in arr if any(xs in s for xs in matchers)]
                        if matching[0] not in filenameArray:
                            filenameArray.append(matching[0])
                        
    print("There are " + str(len(filenameArray)) + " filenames found.")
    print(filenameArray)
    
    outBook = openpyxl.Workbook()
    outSheet = outBook.active
    counter = 1
    for f in filenameArray:
        c = outSheet.cell(row = counter, column = 1)
        c.value = f
        counter += 1
    

    outBook.save("C:\Workspace\Scripts\FilenamesFoundOutput.xlsx")

main()
